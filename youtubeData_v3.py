# -*- coding: utf-8 -*-
"""
Created on Thu Jan 14 21:22:19 2021

@author: Ovindu Wijethunge
"""
from __future__ import unicode_literals
import pickle

import youtube_dl
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from nltk.tokenize import word_tokenize
import re
import numpy as np
from googleapiclient.discovery import build
import speech_recognition as sr
from pydub import AudioSegment
import openpyxl
from pytube import YouTube
from moviepy.editor import *
import os
from main_script_v5 import mainScript

CLIENT_SECRETS_FILE = "client_secret.json"  # for more information   https://python.gotrained.com/youtube-api-extracting-comments/
SCOPES = ['https://www.googleapis.com/auth/youtube.force-ssl']
API_SERVICE_NAME = 'youtube'
API_VERSION = 'v3'

list_date_title = []


def get_video_date(vid):
    if len(list_date_title) != 0:
        list_date_title.clear()

    api_key = 'AIzaSyAMaj2OSJVU_xVC0LyLvmATMDdRgkScIwg'
    youtube = build('youtube', 'v3', developerKey=api_key)
    request = youtube.videos().list(

        part='snippet',
        id=vid
    )
    response = request.execute()
    video_published_date = "date"
    video_title = "helloo"

    for item in response['items']:
        video_published_date = item['snippet']['publishedAt']
        list_date_title.append(video_published_date)
        video_title = item['snippet']['title']
        list_date_title.append(video_title)
    return list_date_title


def get_authenticated_service():
    credentials = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            credentials = pickle.load(token)
    #  Check if the credentials are invalid or do not exist
    if not credentials or not credentials.valid:
        # Check if the credentials have expired
        if credentials and credentials.expired and credentials.refresh_token:
            credentials.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                CLIENT_SECRETS_FILE, SCOPES)
            credentials = flow.run_console()

        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(credentials, token)

    return build(API_SERVICE_NAME, API_VERSION, credentials=credentials)


def write_to_excel_content(content_list):
    wb = openpyxl.Workbook()
    dest_filename = 'contents.xlsx'
    worksheet = wb.active

    worksheet['A1'] = 'content'
    worksheet['B1'] = 'vid'

    row_count = 2
    for listx in content_list:

        list_index = 0
        for j in range(1, len(listx) + 1):
            worksheet.cell(row=row_count, column=j).value = listx[list_index]  # count , j = row , column
            list_index += 1
        row_count = row_count + 1

    wb.save(filename=dest_filename)
    wb.close()


def write_to_excel(d_list):
    wb = openpyxl.Workbook()
    dest_filename = 'comments.xlsx'
    worksheet = wb.active

    worksheet['A1'] = 'vid'
    worksheet['B1'] = 'vdate'
    worksheet['C1'] = 'comment_id'
    worksheet['D1'] = 'comment'
    worksheet['E1'] = 'channel_id'
    worksheet['F1'] = 'video_title'
    worksheet['G1'] = 'like_count'
    worksheet['H1'] = 'published_date'
    worksheet['I1'] = 'updated_date'

    row_count = 2
    for listx in d_list:

        list_index = 0
        for j in range(1, len(listx) + 1):
            worksheet.cell(row=row_count, column=j).value = listx[list_index]  # count , j = row , column
            list_index += 1
        row_count = row_count + 1

    wb.save(filename=dest_filename)
    wb.close()


def convert_audio_to_text(filepath, chunksize = 30000):
    sound = AudioSegment.from_mp3(filepath)

    def divide_chunks(sound, chunksize):
        for i in range(0, len(sound), chunksize):
            yield sound[i:i + chunksize]

    chunks = list(divide_chunks(sound, chunksize))
    print(f"{len(chunks)} chunks of {chunksize / 1000}s each")

    r = sr.Recognizer()
    string_index = {}
    try:
        for index, chunk in enumerate(chunks):
            chunk.export('./sample.wav', format='wav')
            with sr.AudioFile('./sample.wav') as source:
                audio = r.record(source)
            response = r.recognize_google(audio, language="si-LK")

            print("30 second chunk ", response)
            print(index)
            string_index[index] = response
        return string_index
    except sr.UnknownValueError:
        print("sr.UnknownValueError ")
    return string_index


def get_video_content(videoId):
    url = 'https://www.youtube.com/watch?v=' + videoId
    output = "mp3"
    print("Converting...")
    ydl_opts = {
        'format': 'bestaudio/best',
        'postprocessors': [{
            'key': 'FFmpegExtractAudio',
            'preferredcodec': 'mp3',
            'preferredquality': '192',
        }],
    }

    with youtube_dl.YoutubeDL(ydl_opts) as ydl:
        mp4 = ydl.download([url])

    for filename in os.listdir("."):

        if filename[-3:] == 'mp3':
            os.rename(filename, 'sample.mp3')

    text = convert_audio_to_text('sample.mp3')
    os.remove('sample.wav')

    len_of_text = len(text)
    content_string = ""
    for index in range(0, len_of_text):
        content_string = content_string + " " + text[index]

    print("Google speech_recognition text is ", content_string)

    os.remove('sample.mp3')
    return content_string


def de_emojies(text1):
    regrex_pattern = re.compile(pattern="["
                                        u"\U0001F600-\U0001F64F"  # emoticons
                                        u"\U0001F300-\U0001F5FF"  # symbols & pictographs
                                        u"\U0001F680-\U0001F6FF"  # transport & map symbols
                                        u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
                                        u"\U00002702-\U000027B0"
                                        u"\U000024C2-\U0001F251"
                                        "]+", flags=re.UNICODE)
    text = regrex_pattern.sub(r' ', text1)

    return text


def get_video_comments(service, **kwargs):
    data_list = []
    results = service.commentThreads().list(**kwargs).execute()
    while results:

        for item in results['items']:
            try:
                data_row = []
                vid = item['snippet']['topLevelComment']['snippet']['videoId']
                data_row.append(vid)
                vdate = list_date_title[0]
                data_row.append(vdate)
                comment_id = item['snippet']['topLevelComment']['id']
                data_row.append(comment_id)
                comment = item['snippet']['topLevelComment']['snippet']['textDisplay']
                symbols = "!\"#$%&()*+-./:;<=>?@[\]^_`{|}~\n"
                text = comment
                for i in range(len(symbols)):  # for a comment
    
                    text = np.char.replace(text, symbols[i], ' ')
                    # text = np.char.replace(text, "  ", " ")
                    text = np.char.replace(text, ',', ' ')
    
                tokens = word_tokenize(str(text))
                length = len(tokens) / 2
                pattern = re.compile("[A-Za-z]+")
                count = 0
                for token in tokens:  # for a word
                    if pattern.fullmatch(token) is not None:  # if someone use english he does not use sinhala
                        count = count + 1  # detect english words
                if count >= length:
                    continue
                data_row.append(comment)
    
                channel_id = item['snippet']['topLevelComment']['snippet']['authorChannelId']['value']
                data_row.append(channel_id)
    
                video_title = list_date_title[1]
                data_row.append(video_title)
    
                like_count = item['snippet']['topLevelComment']['snippet']['likeCount']
                data_row.append(like_count)
    
                published_date = item['snippet']['topLevelComment']['snippet']['publishedAt']
                data_row.append(published_date)
    
                updated_date = item['snippet']['topLevelComment']['snippet']['updatedAt']
                data_row.append(updated_date)
    
                data_list.append(data_row)
                
                
            except Exception as e:
                 print(e)   
        # Check if another page exists
        if 'nextPageToken' in results:
            kwargs['pageToken'] = results['nextPageToken']
            results = service.commentThreads().list(**kwargs).execute()
        else:
            break

    return data_list


def download_comments_and_content(video_id):
    os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
    service = get_authenticated_service()
    get_video_date(video_id)

    data_list = get_video_comments(service, part='id,snippet', videoId=video_id, textFormat='plainText')

    print(data_list)
    video_content = get_video_content(video_id)  # 0FXKASB1Bd0 _VLjevnS8lw  loO6ws2X50Y# #  BW38guk_fQQ  Rjb9sLL0LZI lUukWG4Fqow
    write_to_excel(data_list)
    content_list = [[video_content,video_id]]
    write_to_excel_content(content_list)
    mainScript()
#idd = 'yRCjfxDAnEU'
#download_comments_and_content(idd)